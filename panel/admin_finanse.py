from os import path

from django.core.exceptions import ValidationError

from .models import (Administrator, Finanse, Budynek)
from simple_history.admin import SimpleHistoryAdmin
from django.contrib import admin
from openpyxl import load_workbook
from django.shortcuts import render, redirect
from django.urls import path
from .forms import ExcelImportForm


class FinanseAdmin(SimpleHistoryAdmin):
	list_display = ['budynek', 'fundusz_remontowy', 'zadluzenie_budynku']
	
	change_list_template = "admin/finanse/change_list.html"
	
	def get_urls(self):
		urls = super().get_urls()
		my_urls = [path('import-excel/', self.admin_site.admin_view(self.import_excel), name='finanse-import-excel')]
		return my_urls + urls
	
	def import_excel(self, request):
		if request.method == "POST":
			form = ExcelImportForm(request.POST, request.FILES)
			if form.is_valid():
				excel_file = form.cleaned_data['excel_file']
				wb = load_workbook(excel_file)
				sheet = wb.active  # Domyślnie pierwsza karta
				
				for row in sheet.iter_rows(min_row=2, values_only=True):  # Min row = 2, aby ominąć nagłówki
					budynek_indeks = row[0]
					fundusz_remontowy = row[2]
					zadluzenie_budynku = row[3]
					
					try:
						budynek = Budynek.objects.get(indeks=budynek_indeks)
						finansowy_budynek, created = Finanse.objects.get_or_create(budynek=budynek)
						finansowy_budynek.fundusz_remontowy = fundusz_remontowy
						finansowy_budynek.zadluzenie_budynku = zadluzenie_budynku
						finansowy_budynek.save()
					
					except Budynek.DoesNotExist:
						self.message_user(request, f"Budynek z indeksem '{budynek_indeks}' nie istnieje.",
										  level='error')
						continue
		
			return render(request, 'admin/success_alert.html')
	
		else:
			form = ExcelImportForm()


		context = {
			'form': form,
			'opts': self.model._meta,
			'app_label': self.model._meta.app_label,
		}
		return render(request, "admin/excel_upload_form.html", context)


def get_queryset(self, request):
	qs = super().get_queryset(request)
	
	if request.user.groups.filter(name='ao_finanse').exists():
		administrator = Administrator.objects.filter(osiedle__nazwa='Ogolna')
		return qs.filter(budynek__administrator__in=administrator)
	
	return qs


admin.site.register(Finanse, FinanseAdmin)
