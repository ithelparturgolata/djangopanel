from os import path
from .models import (Administrator, Energetyka, Budynek)
from simple_history.admin import SimpleHistoryAdmin
from django.urls import path
from django.shortcuts import redirect, render
from django.contrib import admin
from openpyxl import load_workbook
from .forms import ExcelImportForm


class EnergetykaHistoryAdmin(SimpleHistoryAdmin):
	list_display = ['budynek', 'termin_podzielniki', 'wymiana_baterii']
	history_list_display = ['budynek']
	search_fields = ('budynek', 'termin_podzielniki', 'wymiana_baterii')
	
	change_list_template = "admin/energetyka/change_list.html"
	
	def get_urls(self):
		urls = super().get_urls()
		my_urls = [path('import-excel/', self.admin_site.admin_view(self.import_excel), name='energetyka-import-excel'), ]
		return my_urls + urls
	
	def import_excel(self, request):
		if request.method == "POST":
			form = ExcelImportForm(request.POST, request.FILES)
			if form.is_valid():
				excel_file = form.cleaned_data['excel_file']
				wb = load_workbook(excel_file)
				sheet = wb.active
				
				for row in sheet.iter_rows(min_row=2, values_only=True):
					budynek_indeks = row[0]
					termin_podzielniki = row[2] if row[2] is not None else "Brak informacji"
					termin_legalizacji_wodomierzy = row[3] if row[3] is not None else "Brak informacji"
					uwagi_legalizacja_wodomierzy = row[4] if row[4] is not None else "Brak informacji"
					wymiana_baterii = row[5] if row[5] is not None else "Brak informacji"
					uwagi_wymiana_baterii = row[6] if row[6] is not None else "Brak informacji"
					
					try:
						budynek = Budynek.objects.get(indeks=budynek_indeks)
						energetyka_budynek, created = Energetyka.objects.get_or_create(budynek=budynek)
						energetyka_budynek.termin_podzielniki = termin_podzielniki
						energetyka_budynek.termin_legalizacji_wodomierzy = termin_legalizacji_wodomierzy
						energetyka_budynek.uwagi_legalizacja_wodomierzy = uwagi_legalizacja_wodomierzy
						energetyka_budynek.wymiana_baterii = wymiana_baterii
						energetyka_budynek.uwagi_wymiana_baterii = uwagi_wymiana_baterii
						energetyka_budynek.save()
						
					except Budynek.DoesNotExist:
						self.message_user(request, f"Budynek z indeksem '{budynek_indeks}' nie istnieje.",
										  level='error')
						continue
				
			return render(request, 'admin/success_alert.html')
		else:
			form = ExcelImportForm()
		
		context = {
			'form': form,
			'opts': self.model._meta,
			'app_label': self.model._meta.app_label,
		}
		return render(request, "admin/excel_upload_form.html", context)
	
	def get_queryset(self, request):
		qs = super().get_queryset(request)
		
		if request.user.groups.filter(name='ao_finanse').exists():
			administrator = Administrator.objects.filter(osiedle__nazwa='Ogolna')
			return qs.filter(budynek__administrator__in=administrator)
		
		return qs


admin.site.register(Energetyka, EnergetykaHistoryAdmin)
