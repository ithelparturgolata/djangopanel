from os import path
from .models import (Administrator, Przeglady, Budynek)
from simple_history.admin import SimpleHistoryAdmin
from django.contrib import admin
from openpyxl import load_workbook
from django.shortcuts import render, redirect
from django.urls import path
from .forms import ExcelImportForm


class PrzegladyHistoryAdmin(SimpleHistoryAdmin):
    list_display = ['budynek']
    history_list_display = ['nazwa']
    search_fields = ['budynek']
    
    change_list_template = "admin/przeglady/change_list.html"
    
    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel), name='przeglady-import-excel'), ]
        return my_urls + urls
    
    def import_excel(self, request):
        if request.method == "POST":
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = form.cleaned_data['excel_file']
                wb = load_workbook(excel_file)
                sheet = wb.active
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    budynek_indeks = row[0]
                    planowany_przeglad_kominiarski = row[2] if row[2] is not None else "Brak informacji"
                    planowany_przeglad_gazowy = row[3] if row[3] is not None else "Brak informacji"
                    planowany_przeglad_roczny_budynku = row[4] if row[4] is not None else "Brak informacji"
                    planowany_przeglad_piecioletni_budynku = row[5] if row[5] is not None else "Brak informacji"
                    planowany_przeglad_piecioletni_gazowy = row[6] if row[6] is not None else "Brak informacji"
                    aktualny_przeglad_kominiarski = row[7] if row[7] is not None else "Brak informacji"
                    aktualny_przeglad_gazowy = row[8] if row[8] is not None else "Brak informacji"
                    aktualny_przeglad_roczny_budynku = row[9] if row[9] is not None else "Brak informacji"
                    aktualny_przeglad_piecioletni_budynku = row[10] if row[10] is not None else "Brak informacji"
                    aktualny_przeglad_piecioletni_gazowy = row[11] if row[11] is not None else "Brak informacji"
                    
                    try:
                        budynek = Budynek.objects.get(indeks=budynek_indeks)
                        przeglady_budynek, created = Przeglady.objects.get_or_create(budynek=budynek)
                        przeglady_budynek.planowany_przeglad_kominiarski = planowany_przeglad_kominiarski
                        przeglady_budynek.planowany_przeglad_gazowy = planowany_przeglad_gazowy
                        przeglady_budynek.planowany_przeglad_roczny_budynku = planowany_przeglad_roczny_budynku
                        przeglady_budynek.planowany_przeglad_piecioletni_budynku = planowany_przeglad_piecioletni_budynku
                        przeglady_budynek.planowany_przeglad_piecioletni_gazowy = planowany_przeglad_piecioletni_gazowy
                        przeglady_budynek.aktualny_przeglad_kominiarski = aktualny_przeglad_kominiarski
                        przeglady_budynek.aktualny_przeglad_gazowy = aktualny_przeglad_gazowy
                        przeglady_budynek.aktualny_przeglad_roczny_budynku = aktualny_przeglad_roczny_budynku
                        przeglady_budynek.aktualny_przeglad_piecioletni_budynku = aktualny_przeglad_piecioletni_budynku
                        przeglady_budynek.aktualny_przeglad_piecioletni_gazowy = aktualny_przeglad_piecioletni_gazowy
                        przeglady_budynek.save()
                    except Budynek.DoesNotExist:
                        self.message_user(request, f"Budynek z indeksem '{budynek_indeks}' nie istnieje.",
                                          level='error')
                        continue
                
                return render(request, 'admin/success_alert.html')
        else:
            form = ExcelImportForm()
        
        context = {
            'form': form,
            'opts': self.model._meta,
            'app_label': self.model._meta.app_label,
        }
        return render(request, "admin/excel_upload_form.html", context)
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        
        if request.user.groups.filter(name='ns_przeglady').exists():
            administrator = Administrator.objects.filter(osiedle__nazwa='Na Skarpie')
            return qs.filter(budynek__administrator__in=administrator)
        elif request.user.groups.filter(name='ce_przeglady').exists():
            administrator = Administrator.objects.filter(osiedle__nazwa='Centrum')
            return qs.filter(budynek__administrator__in=administrator)
        elif request.user.groups.filter(name='nw_przeglady').exists():
            administrator = Administrator.objects.filter(osiedle__nazwa='Na Wyżynach')
            return qs.filter(budynek__administrator__in=administrator)
        
        return qs


admin.site.register(Przeglady, PrzegladyHistoryAdmin)

