from os import path
from .models import (Administrator, Smieci, Budynek)
from simple_history.admin import SimpleHistoryAdmin
from django.urls import path
from django.shortcuts import render
from django.contrib import admin
from openpyxl import load_workbook
from .forms import ExcelImportForm


class SmieciAdmin(SimpleHistoryAdmin):
	list_display = ['budynek', 'styczen', 'luty', 'marzec', 'kwiecien', 'maj', 'czerwiec', 'lipiec',
					'sierpien', 'wrzesien', 'pazdziernik', 'listopad', 'grudzien']
	
	change_list_template = "admin/smieci/change_list.html"
	
	def get_urls(self):
		urls = super().get_urls()
		my_urls = [
			path('import-excel/', self.admin_site.admin_view(self.import_excel), name='smieci-import-excel'),
		]
		return my_urls + urls
	
	def import_excel(self, request):
		if request.method == "POST":
			form = ExcelImportForm(request.POST, request.FILES)
			if form.is_valid():
				excel_file = form.cleaned_data['excel_file']
				wb = load_workbook(excel_file)
				sheet = wb.active
				
				for row in sheet.iter_rows(min_row=2, values_only=True):
					budynek_indeks = row[0]
					styczen = row[2] if row[2] is not None else "Brak informacji"
					luty = row[3] if row[3] is not None else "Brak informacji"
					marzec = row[4] if row[4] is not None else "Brak informacji"
					kwiecien = row[5] if row[5] is not None else "Brak informacji"
					maj = row[6] if row[6] is not None else "Brak informacji"
					czerwiec = row[7] if row[7] is not None else "Brak informacji"
					lipiec = row[8] if row[8] is not None else "Brak informacji"
					sierpien = row[9] if row[9] is not None else "Brak informacji"
					wrzesien = row[10] if row[10] is not None else "Brak informacji"
					pazdziernik = row[11] if row[11] is not None else "Brak informacji"
					listopad = row[12] if row[12] is not None else "Brak informacji"
					grudzien = row[13] if row[13] is not None else "Brak informacji"
					
					try:
						budynek = Budynek.objects.get(indeks=budynek_indeks)
						smieci_budynek, created = Smieci.objects.get_or_create(budynek=budynek)
						smieci_budynek.styczen = styczen
						smieci_budynek.luty = luty
						smieci_budynek.marzec = marzec
						smieci_budynek.kwiecien = kwiecien
						smieci_budynek.maj = maj
						smieci_budynek.czerwiec = czerwiec
						smieci_budynek.lipiec = lipiec
						smieci_budynek.sierpien = sierpien
						smieci_budynek.wrzesien = wrzesien
						smieci_budynek.pazdziernik = pazdziernik
						smieci_budynek.listopad = listopad
						smieci_budynek.grudzien = grudzien
						smieci_budynek.save()
					except Budynek.DoesNotExist:
						self.message_user(request, f"Budynek z indeksem '{budynek_indeks}' nie istnieje.",
										  level='error')
						continue
			
			return render(request, 'admin/success_alert.html')
		else:
			form = ExcelImportForm()
		
		context = {
			'form': form,
			'opts': self.model._meta,
			'app_label': self.model._meta.app_label,
		}
		return render(request, "admin/excel_upload_form.html", context)
	
	def get_queryset(self, request):
		qs = super().get_queryset(request)
		
		if request.user.groups.filter(name='ns_smieci').exists():
			administrator = Administrator.objects.filter(osiedle__nazwa='Na Skarpie')
			return qs.filter(budynek__administrator__in=administrator)
		elif request.user.groups.filter(name='ce_smieci').exists():
			administrator = Administrator.objects.filter(osiedle__nazwa='Centrum')
			return qs.filter(budynek__administrator__in=administrator)
		elif request.user.groups.filter(name='nw_smieci').exists():
			administrator = Administrator.objects.filter(osiedle__nazwa='Na Wyżynach')
			return qs.filter(budynek__administrator__in=administrator)
		
		return qs


admin.site.register(Smieci, SmieciAdmin)
