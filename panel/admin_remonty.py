from os import path

from django.utils.html import format_html
from django.utils.timezone import now

from .models import (Administrator, Finanse, Budynek, Remonty)
from simple_history.admin import SimpleHistoryAdmin
from django.contrib import admin
from openpyxl import load_workbook
from django.shortcuts import render, redirect
from django.urls import path
from .forms import ExcelImportForm


class RemontyAdmin(SimpleHistoryAdmin):
	list_display = ('budynek', 'wykonawca', 'koniec_prac', 'status_remontu')
	history_list_display = ['budynek', 'wykonawca']
	search_fields = ('ulica', 'indeks', 'wykonawca', 'status_remontu')
	
	def status_remontu(self, obj):
		if obj.koniec_prac and obj.koniec_prac < now().date():
			return format_html('<span style="color: red;">Po terminie</span>')
		return format_html('<span style="color: green;">W trakcie</span>')
	
	status_remontu.short_description = 'Status Remontu'
	
	change_list_template = "admin/remonty/change_list.html"
	
	def get_urls(self):
		urls = super().get_urls()
		my_urls = [
			path('import-excel/', self.admin_site.admin_view(self.import_excel), name='remonty-import-excel'), ]
		return my_urls + urls
	
	def import_excel(self, request):
		if request.method == "POST":
			form = ExcelImportForm(request.POST, request.FILES)
			if form.is_valid():
				excel_file = form.cleaned_data['excel_file']
				wb = load_workbook(excel_file)
				sheet = wb.active
				
				for row in sheet.iter_rows(min_row=2, values_only=True):
					budynek_indeks = row[0]
					wykonawca = row[2]
					opis_remontu = row[3]
					poczatek_prac = row[4]
					koniec_prac = row[5]
					
					try:
						budynek = Budynek.objects.get(indeks=budynek_indeks)
						remonty_budynek, created = Remonty.objects.get_or_create(budynek=budynek)
						remonty_budynek.wykonawca = wykonawca
						remonty_budynek.opis_remontu = opis_remontu
						remonty_budynek.poczatek_prac = poczatek_prac
						remonty_budynek.koniec_prac = koniec_prac
						
						remonty_budynek.save()
					except Budynek.DoesNotExist:
						self.message_user(request, f"Budynek z indeksem '{budynek_indeks}' nie istnieje.",
										  level='error')
						continue
			
			return render(request, 'admin/success_alert.html')
		else:
			form = ExcelImportForm()
		
		context = {
			'form': form,
			'opts': self.model._meta,
			'app_label': self.model._meta.app_label,
		}
		return render(request, "admin/excel_upload_form.html", context)
	
	def get_queryset(self, request):
		qs = super().get_queryset(request)
		
		if request.user.groups.filter(name='ns_remonty').exists():
			administrator = Administrator.objects.filter(osiedle__nazwa='Na Skarpie')
			return qs.filter(budynek__administrator__in=administrator)
		elif request.user.groups.filter(name='ce_remonty').exists():
			administrator = Administrator.objects.filter(osiedle__nazwa='Centrum')
			return qs.filter(budynek__administrator__in=administrator)
		elif request.user.groups.filter(name='nw_remonty').exists():
			administrator = Administrator.objects.filter(osiedle__nazwa='Na Wyżynach')
			return qs.filter(budynek__administrator__in=administrator)
		
		return qs


admin.site.register(Remonty, RemontyAdmin)
