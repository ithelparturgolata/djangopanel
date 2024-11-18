from os import path
from .models import Budynek, Administrator
from simple_history.admin import SimpleHistoryAdmin
from django.contrib import admin
from openpyxl import load_workbook
from django.shortcuts import render, redirect
from django.urls import path
from .forms import ExcelImportForm


class BudynekAdmin(SimpleHistoryAdmin):
	list_display = ['ulica', 'indeks']
	search_fields = ['indeks', 'ulica']
	
	change_list_template = 'admin/budynek/change_list.html'
	
	def get_urls(self):
		urls = super().get_urls()
		my_urls = [path('import-excel/', self.admin_site.admin_view(self.import_excel), name='budynek-import-excel'), ]
		return my_urls + urls
	
	def import_excel(self, request):
		if request.method == "POST":
			form = ExcelImportForm(request.POST, request.FILES)
			if form.is_valid():
				excel_file = form.cleaned_data['excel_file']
				wb = load_workbook(excel_file)
				sheet = wb.active
				
				for row in sheet.iter_rows(min_row=2, values_only=True):
					budynek_indeks = row[0]
					ulica = row[1]
					kod = row[2]
					rok_budowy = row[3]
					telefon_osiedla = row[4]
					obreb = row[5]
					dzialka = row[6]
					laczna_powierzchnia_dzialki = row[7]
					powierzchnia_wspolna_budynku = row[8]
					kierownik_id = row[9]
					osiedle_id = row[10]
					spoldzielnia_id = row[11]
					administrator_id = row[12]
					
					try:
						budynek, created = Budynek.objects.get_or_create(
							indeks=budynek_indeks,
							defaults={
								'ulica': ulica,
								'kod': kod,
								'rok_budowy': rok_budowy,
								'telefon_osiedla': telefon_osiedla,
								'obreb': obreb,
								'dzialka': dzialka,
								'laczna_powierzchnia_dzialki': laczna_powierzchnia_dzialki,
								'powierzchnia_wspolna_budynku': powierzchnia_wspolna_budynku,
								'kierownik_id' : kierownik_id,
								'osiedle_id': osiedle_id,
								'spoldzielnia_id': spoldzielnia_id,
								'administrator_id': administrator_id
							}
						)
						if not created:
							budynek.ulica = ulica
							budynek.kod = kod
							budynek.rok_budowy = rok_budowy
							budynek.telefon_osiedla = telefon_osiedla
							budynek.obreb = obreb
							budynek.dzialka = dzialka
							budynek.laczna_powierzchnia_dzialki = laczna_powierzchnia_dzialki
							budynek.powierzchnia_wspolna_budynku = powierzchnia_wspolna_budynku
							budynek.kierownik_id = kierownik_id
							budynek.osiedle_id = osiedle_id
							budynek.spoldzielnia_id = spoldzielnia_id
							budynek.administrator_id = administrator_id
							budynek.save()
					
					except Exception as e:
						self.message_user(request, f"Błąd podczas przetwarzania budynku '{budynek_indeks}': {str(e)}",
										  level='error')
						continue
				
				self.message_user(request, "Dane z pliku Excel zostały pomyślnie zaktualizowane.")
				return redirect('admin:budynek-import-excel')
		else:
			form = ExcelImportForm()
		
		context = {
			'form': form,
			'opts': self.model._meta,
			'app_label': self.model._meta.app_label,
		}
		return render(request, "admin/excel_upload_form.html", context)
	
	def get_queryset(self, request):
		qs = super().get_queryset(request)
		
		if request.user.groups.filter(name='ns_budynki').exists():
			administrator = Administrator.objects.filter(osiedle__nazwa='Na Skarpie')
			return qs.filter(budynek__administrator__in=administrator)
		elif request.user.groups.filter(name='ce_budynki').exists():
			administrator = Administrator.objects.filter(osiedle__nazwa='Centrum')
			return qs.filter(budynek__administrator__in=administrator)
		elif request.user.groups.filter(name='nw_budynki').exists():
			administrator = Administrator.objects.filter(osiedle__nazwa='Na Wyżynach')
			return qs.filter(budynek__administrator__in=administrator)
		
		return qs


admin.site.register(Budynek, BudynekAdmin)
